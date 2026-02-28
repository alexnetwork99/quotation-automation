from dotenv import load_dotenv
load_dotenv()

import os
import re
import json
from pathlib import Path

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.security import APIKeyHeader
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
import chromadb
from zhipuai import ZhipuAI

# ── Config ────────────────────────────────────────────────────────────────────
ZHIPU_API_KEY = os.environ.get("ZHIPU_API_KEY", "")
GLM_MODEL = "glm-4-flash"
EMBED_MODEL = "embedding-3"
DASHSCOPE_API_KEY = os.environ.get("DASHSCOPE_API_KEY", "")
CHROMA_PATH = Path(__file__).parent / "chroma_db"
DATA_PATH = Path(__file__).parent
FRONTEND_PATH = Path(__file__).parent / "frontend" / "index.html"

API_KEY = "edwardluo"
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

async def verify_token(key: str = Depends(api_key_header)):
    if key != API_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return key

app = FastAPI()

def strip_json(raw: str) -> str:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
    return raw.strip()
zhipu = ZhipuAI(api_key=ZHIPU_API_KEY)
chroma = chromadb.PersistentClient(path=str(CHROMA_PATH))
collection = chroma.get_or_create_collection("price_library")


# ── Embedding ─────────────────────────────────────────────────────────────────
def embed(text: str) -> list:
    resp = zhipu.embeddings.create(model=EMBED_MODEL, input=text)
    return resp.data[0].embedding

# ── 解析价格库 txt ─────────────────────────────────────────────────────────────
def parse_price_file(filepath: Path) -> list[dict]:
    items = []
    supplier = ""
    with open(filepath, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("供应商："):
                supplier = line[4:]
            m = re.match(r"品名：(.+?)，规格：(.+?)，单位：(.+?)，单价：(.+?)元", line)
            if m:
                items.append({
                    "supplier": supplier,
                    "name": m.group(1),
                    "spec": m.group(2),
                    "unit": m.group(3),
                    "price": float(m.group(4)),
                })
    return items

# ── 初始化：导入价格库到 ChromaDB ──────────────────────────────────────────────
def init_price_library():
    if collection.count() > 0:
        return
    txt_files = list(DATA_PATH.glob("*.txt"))
    for f in txt_files:
        items = parse_price_file(f)
        for i, item in enumerate(items):
            text = f"{item['name']} {item['spec']} 供应商:{item['supplier']}"
            doc_id = f"{f.stem}_{i}"
            collection.add(
                ids=[doc_id],
                embeddings=[embed(text)],
                documents=[text],
                metadatas=[item],
            )

init_price_library()


# ── Models ────────────────────────────────────────────────────────────────────
class QuoteRequest(BaseModel):
    inquiry: str
    margin: float = 0.25  # 利润率，默认 25%

class PriceItem(BaseModel):
    supplier: str
    name: str
    spec: str
    unit: str
    price: float

# ── Tab1：报价生成 ─────────────────────────────────────────────────────────────
@app.post("/scene1/api/quote")
async def generate_quote(req: QuoteRequest, _=Depends(verify_token)):
    query_vec = embed(req.inquiry)
    results = collection.query(query_embeddings=[query_vec], n_results=5)
    if not results["metadatas"][0]:
        raise HTTPException(status_code=404, detail="未找到匹配产品")

    candidates = []
    for meta, doc in zip(results["metadatas"][0], results["documents"][0]):
        candidates.append(
            f"品名:{meta['name']} 规格:{meta['spec']} 单价:{meta['price']}元/{meta['unit']} 供应商:{meta['supplier']}"
        )
    candidates_text = "\n".join(candidates)

    prompt = f"""你是一个外贸五金厂的报价助手。根据客户询价和候选产品，生成一份报价单。

客户询价：{req.inquiry}

候选产品（成本价）：
{candidates_text}

利润率：{int(req.margin * 100)}%
客户报价 = 成本价 × (1 + 利润率)

请从候选产品中选择最匹配的产品，生成报价单。返回 JSON 格式：
{{
  "items": [
    {{"name": "产品名", "spec": "规格", "unit": "单位", "cost_price": 成本单价, "unit_price": 客户单价, "quantity": 数量, "total": 客户小计, "supplier": "供应商"}}
  ],
  "total_amount": 客户总金额,
  "total_cost": 总成本,
  "note": "备注"
}}
JSON only, no markdown."""

    resp = zhipu.chat.completions.create(
        model=GLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = strip_json(resp.choices[0].message.content)
    try:
        data = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=500, detail=f"LLM 返回格式错误: {raw}")
    return data


# ── Tab2：价格库管理 ───────────────────────────────────────────────────────────
@app.get("/scene1/api/prices")
async def list_prices(_=Depends(verify_token)):
    results = collection.get(include=["metadatas"])
    items = []
    for doc_id, meta in zip(results["ids"], results["metadatas"]):
        items.append({"id": doc_id, **meta})
    return items

@app.post("/scene1/api/prices")
async def add_price(item: PriceItem, _=Depends(verify_token)):
    import uuid
    doc_id = str(uuid.uuid4())
    text = f"{item.name} {item.spec} 供应商:{item.supplier}"
    collection.add(
        ids=[doc_id],
        embeddings=[embed(text)],
        documents=[text],
        metadatas=[item.model_dump()],
    )
    return {"id": doc_id, "status": "ok"}

@app.delete("/scene1/api/prices/{doc_id}")
async def delete_price(doc_id: str, _=Depends(verify_token)):
    collection.delete(ids=[doc_id])
    return {"status": "ok"}

@app.delete("/scene1/api/prices", dependencies=[Depends(verify_token)])
async def clear_prices():
    ids = collection.get(include=[])["ids"]
    if ids:
        collection.delete(ids=ids)
    return {"status": "ok", "deleted": len(ids)}

class ImportTextRequest(BaseModel):
    text: str

@app.post("/scene1/api/prices/import-text")
async def import_text(req: ImportTextRequest, _=Depends(verify_token)):
    prompt = f"""从以下文本中提取产品价格信息，返回 JSON 数组，每项包含：supplier（供应商）、name（品名）、spec（规格）、unit（单位）、price（单价，数字）。
若某字段无法确定，supplier 默认"未知供应商"，spec 默认"-"，unit 默认"个"。
只返回 JSON 数组，不要 markdown。

文本：
{req.text}"""
    resp = zhipu.chat.completions.create(
        model=GLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = strip_json(resp.choices[0].message.content)
    try:
        items = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=500, detail=f"AI 解析失败: {raw[:200]}")
    return await _bulk_insert(items)

@app.post("/scene1/api/prices/import-file")
async def import_file(file: UploadFile = File(...), _=Depends(verify_token)):
    content = await file.read()
    filename = file.filename or ""
    items = []
    if filename.endswith(".txt"):
        import tempfile, os
        with tempfile.NamedTemporaryFile(delete=False, suffix=".txt") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        items = parse_price_file(Path(tmp_path))
        os.unlink(tmp_path)
    elif filename.endswith((".xlsx", ".xls")):
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                items.append({
                    "supplier": str(row[col.get("供应商", 0)] or "未知供应商"),
                    "name": str(row[col.get("品名", 1)] or ""),
                    "spec": str(row[col.get("规格", 2)] or "-"),
                    "unit": str(row[col.get("单位", 3)] or "个"),
                    "price": float(row[col.get("单价", 4)] or 0),
                })
            except Exception:
                continue
    else:
        raise HTTPException(status_code=400, detail="仅支持 .txt / .xlsx 文件")
    return await _bulk_insert(items)

class DeleteQueryRequest(BaseModel):
    query: str

@app.post("/scene1/api/prices/delete-by-query")
async def delete_by_query(req: DeleteQueryRequest, _=Depends(verify_token)):
    """AI 解析自然语言，返回匹配条目供前端预览（不直接删除）"""
    all_items = collection.get(include=["metadatas"])
    if not all_items["ids"]:
        return {"matches": []}
    items_text = "\n".join(
        f"id:{doc_id} 品名:{m['name']} 规格:{m['spec']} 单价:{m['price']} 供应商:{m['supplier']}"
        for doc_id, m in zip(all_items["ids"], all_items["metadatas"])
    )
    prompt = f"""价格库条目列表：
{items_text}

用户指令：{req.query}

请找出符合用户指令的条目，返回 JSON 数组，每项包含 id 字段。只返回 JSON 数组，不要 markdown。"""
    resp = zhipu.chat.completions.create(
        model=GLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = strip_json(resp.choices[0].message.content)
    try:
        matched_ids = [item["id"] for item in json.loads(raw)]
    except Exception:
        raise HTTPException(status_code=500, detail=f"AI 解析失败: {raw[:200]}")
    matches = [
        {"id": doc_id, **meta}
        for doc_id, meta in zip(all_items["ids"], all_items["metadatas"])
        if doc_id in matched_ids
    ]
    return {"matches": matches}

@app.post("/scene1/api/prices/delete-confirm")
async def delete_confirm(body: dict, _=Depends(verify_token)):
    """确认删除指定 id 列表"""
    ids = body.get("ids", [])
    if ids:
        collection.delete(ids=ids)
    return {"deleted": len(ids)}

async def _bulk_insert(items: list) -> dict:
    import uuid
    ok, fail = 0, 0
    for item in items:
        try:
            doc_id = str(uuid.uuid4())
            text = f"{item['name']} {item['spec']} 供应商:{item['supplier']}"
            collection.add(
                ids=[doc_id],
                embeddings=[embed(text)],
                documents=[text],
                metadatas=[{"supplier": str(item["supplier"]), "name": str(item["name"]),
                            "spec": str(item["spec"]), "unit": str(item["unit"]),
                            "price": float(item["price"])}],
            )
            ok += 1
        except Exception:
            fail += 1
    return {"imported": ok, "failed": fail}




# ── 前端 HTML ─────────────────────────────────────────────────────────────────
@app.get("/scene1", response_class=HTMLResponse)
@app.get("/", response_class=HTMLResponse)
async def index():
    return HTMLResponse(FRONTEND_PATH.read_text(encoding="utf-8"))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8081)
